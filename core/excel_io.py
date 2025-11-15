from typing import List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from .translations import get_translation


def create_excel_table(
    ws: Worksheet,
    start_row: int,
    start_col: int,
    end_row: int,
    end_col: int,
    table_name: str,
) -> Table:
    """Create a named table in openpyxl covering the given range"""
    start_col_letter = get_column_letter(start_col)
    end_col_letter = get_column_letter(end_col)
    ref = f"{start_col_letter}{start_row}:{end_col_letter}{end_row}"
    table = Table(displayName=table_name, ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)
    return table


def write_rules_sheet_openpyxl(
    wb: Workbook, rules: List[Tuple[str, str]], language: str = "en"
):
    if "Rules" in wb.sheetnames:
        return wb["Rules"]

    ws: Worksheet = wb.create_sheet("Rules")
    ws.append(
        [get_translation("pattern", language), get_translation("category", language)]
    )
    for pat, cat in rules:
        ws.append([pat, cat])

    create_excel_table(
        ws,
        start_row=1,
        start_col=1,
        end_row=len(rules) + 1,
        end_col=2,
        table_name="Rules",
    )
    # make header bold
    for cell in ws[1]:
        cell.font = Font(bold=True)
    return ws


def write_transactions_sheet_openpyxl(
    wb: Workbook,
    sheet_name: str,
    columns: List[tuple],
    transactions: List[object],
    rules: List[Tuple[str, str]],
    language: str = "en",
):
    if sheet_name in wb.sheetnames:
        raise ValueError(f"Sheet {sheet_name} already exists")

    ws: Worksheet = wb.create_sheet(sheet_name)
    # write headers from columns (each column is (key, label))
    headers = [label for (_, label) in columns]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    start_row = 2
    for i, tx in enumerate(transactions, start=start_row):
        row = []
        for key, _ in columns:
            # prefer attribute access, then dict key
            val = getattr(tx, key, None)
            if val is None and isinstance(tx, dict):
                val = tx.get(key)
            row.append(val)
        ws.append(row)

        # category formulas will be written after table creation using structured references

    # create table - reuse existing table if it matches same sheet+range; otherwise
    # create a unique table name derived from the sheet name and a short hash to avoid collisions
    end_row = start_row + len(transactions) - 1 if transactions else start_row

    table: Table = create_excel_table(
        ws,
        start_row=1,
        start_col=1,
        end_row=end_row,
        end_col=len(headers),
        table_name=sheet_name + "_transactions",
    )

    # After table creation, set a calculated column formula for the category column (if rules present)
    try:
        if rules:
            rules_category_label = get_translation("category", language)
            store_label = get_translation("store", language)
            pattern_label = get_translation("pattern", language)
            category_label = rules_category_label

            col_names = [lab for (_k, lab) in columns]

            cat_index = col_names.index(category_label)
            if cat_index is not None:
                formula = f"=INDEX(Rules[{rules_category_label}],MATCH(1,INDEX(--ISNUMBER(SEARCH(INDEX(Rules[{pattern_label}],0),@[{store_label}])),0),0))"

                try:
                    table.tableColumns[cat_index].calculatedColumnFormula = formula
                except Exception:
                    # Some openpyxl versions or table states may not allow assignment; fall back to per-cell formulas
                    for r in range(2, end_row + 1):
                        ws.cell(row=r, column=cat_index + 1).value = formula
    except Exception:
        # If anything fails, don't block writing the workbook
        pass
    return ws


def write_summary_section_openpyxl(
    ws: Worksheet, summary: List[dict], start_col: int, language: str = "en"
):
    headers = [
        get_translation("over_x_months", language),
        get_translation("sum", language),
    ]
    start_row = 2
    col = start_col
    # write header row for summary table
    for j, h in enumerate(headers, start=col):
        ws.cell(row=start_row - 1, column=j, value=h).font = Font(bold=True)

    for i, row in enumerate(summary, start=start_row):
        ws.cell(row=i, column=col, value=row["months"])
        ws.cell(row=i, column=col + 1, value=row["sum"])

    end_row = start_row + len(summary) - 1 if summary else start_row
    end_col = col + len(headers) - 1

    srow = start_row - 1
    scol = col
    erow = end_row
    ecol = end_col
    base_name = ws.title + "_summary"

    create_excel_table(
        ws,
        start_row=srow,
        start_col=scol,
        end_row=erow,
        end_col=ecol,
        table_name=base_name,
    )
    return ws
